"""
Fill the DTA3 Excel template with combined day tour + overnight arrival data.

Overnight total visitors for the Excel = overnight_nights × visitor_count
(nights stayed × number of visitors), but this multiplied figure is only
used in the exported file — the dashboard always shows raw counts.
"""

from __future__ import annotations

import io
import re
from calendar import month_name
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from services.arrival_reports import monthly_spot_reports_for_export
from services.lgus import list_lgus_simple

_STATIC = Path(__file__).resolve().parent.parent / "static" / "excel"
DTA3_TEMPLATE = _STATIC / "DTA3 (VAR2).xlsx"

_COUNT_KEYS = (
    "this_city_male",
    "this_city_female",
    "other_city_male",
    "other_city_female",
    "other_province_male",
    "other_province_female",
    "foreign_male",
    "foreign_female",
)


def _parse_report_date(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        return date.fromisoformat(value[:10])
    return date.today()


def _lgu_name(row: dict[str, Any]) -> str:
    lgu = row.get("lgus") or {}
    if isinstance(lgu, dict) and lgu.get("name"):
        return str(lgu["name"])
    return ""


def _spot_name(row: dict[str, Any]) -> str:
    spot = row.get("tourist_spots") or {}
    if isinstance(spot, dict) and spot.get("name"):
        return str(spot["name"])
    return f"Spot #{row.get('tourist_spot_id', '')}"


def _spot_code(row: dict[str, Any]) -> int | str | None:
    spot = row.get("tourist_spots") or {}
    if isinstance(spot, dict) and spot.get("code") is not None:
        return spot["code"]
    return None


def _month_year_label(report_date: date) -> str:
    return f"{month_name[report_date.month]} {report_date.year}"


def _safe_filename_part(name: str) -> str:
    return re.sub(r"[^\w\-]+", "_", name.strip()).strip("_") or "LGU"


def _lgu_label(lgu_id: int) -> str:
    for row in list_lgus_simple():
        if int(row["id"]) == lgu_id:
            return str(row["name"])
    return f"LGU_{lgu_id}"


def _overnight_excel_visitors(rep: dict[str, Any]) -> int:
    """
    For the Excel export only: overnight total = night_stay × visitor_count.
    visitor_count is the sum of the 8 origin/sex fields.
    """
    visitor_count = sum(int(rep.get(k) or 0) for k in _COUNT_KEYS)
    nights = int(rep.get("overnight_nights") or 0)
    return nights * visitor_count


def _scale_counts(rep: dict[str, Any]) -> dict[str, int]:
    """
    Return count fields scaled by overnight_nights for the Excel export.
    Each origin/sex cell = raw_count × nights.
    """
    nights = int(rep.get("overnight_nights") or 0)
    return {k: int(rep.get(k) or 0) * nights for k in _COUNT_KEYS}


def _merge_by_spot(
    day_reports: list[dict[str, Any]],
    night_reports: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """
    Merge day tour and overnight reports into one row per tourist spot.

    Day tour counts are used as-is.
    Overnight counts are scaled by overnight_nights before merging
    (night_stay × visitors per origin/sex cell).

    The merged rows are sorted by spot name for consistent output.
    """
    merged: dict[int, dict[str, Any]] = {}

    for rep in day_reports:
        sid = rep.get("tourist_spot_id")
        if sid is None:
            continue
        sid = int(sid)
        if sid not in merged:
            merged[sid] = {
                "tourist_spot_id": sid,
                "_spot_name": _spot_name(rep),
                "_spot_code": _spot_code(rep),
                "_lgu_name": _lgu_name(rep),
                "_report_date": rep.get("report_date"),
                **{k: 0 for k in _COUNT_KEYS},
            }
        for k in _COUNT_KEYS:
            merged[sid][k] += int(rep.get(k) or 0)

    for rep in night_reports:
        sid = rep.get("tourist_spot_id")
        if sid is None:
            continue
        sid = int(sid)
        scaled = _scale_counts(rep)
        if sid not in merged:
            merged[sid] = {
                "tourist_spot_id": sid,
                "_spot_name": _spot_name(rep),
                "_spot_code": _spot_code(rep),
                "_lgu_name": _lgu_name(rep),
                "_report_date": rep.get("report_date"),
                **{k: 0 for k in _COUNT_KEYS},
            }
        for k in _COUNT_KEYS:
            merged[sid][k] += scaled[k]

    return sorted(merged.values(), key=lambda x: x["_spot_name"])


def export_combined_workbook(
    *,
    lgu_id: int,
    report_date: date | None = None,
) -> tuple[bytes, str]:
    """
    Export a single DTA3 workbook for one LGU.

    Day tour and overnight data are merged into ONE row per tourist spot —
    overnight counts are scaled (night_stay × visitors) before summing.
    The totals row reflects the combined grand total.
    """
    day_reports = monthly_spot_reports_for_export(
        lgu_id=lgu_id,
        report_date=report_date,
        visitor_category="day_tour",
    )
    if not day_reports and report_date:
        day_reports = monthly_spot_reports_for_export(
            lgu_id=lgu_id, visitor_category="day_tour"
        )

    night_reports = monthly_spot_reports_for_export(
        lgu_id=lgu_id,
        report_date=report_date,
        visitor_category="overnight",
    )
    if not night_reports and report_date:
        night_reports = monthly_spot_reports_for_export(
            lgu_id=lgu_id, visitor_category="overnight"
        )

    lgu_title = _lgu_label(lgu_id)
    wb = openpyxl.load_workbook(DTA3_TEMPLATE)
    ws = wb.active

    # Header cells — use whichever dataset has data
    all_reports = day_reports + night_reports
    if all_reports:
        rd = _parse_report_date(all_reports[0].get("report_date"))
        ws["F5"] = _month_year_label(rd)
        first = all_reports[0]
        ws["F6"] = _lgu_name(first) or lgu_title
    else:
        ws["F6"] = lgu_title

    # Merge both datasets into one row per spot
    rows = _merge_by_spot(day_reports, night_reports)

    start_row = 12
    for i, row in enumerate(rows[:50]):
        r = start_row + i
        ws.cell(r, 2, row["_spot_name"])
        if row["_spot_code"] is not None:
            ws.cell(r, 3, row["_spot_code"])
        ws.cell(r, 4,  row["this_city_male"])
        ws.cell(r, 5,  row["this_city_female"])
        ws.cell(r, 7,  row["other_city_male"])
        ws.cell(r, 8,  row["other_city_female"])
        ws.cell(r, 10, row["other_province_male"])
        ws.cell(r, 11, row["other_province_female"])
        ws.cell(r, 13, row["foreign_male"])
        if ws.max_column >= 14:
            ws.cell(r, 14, row["foreign_female"])

    # Grand totals row
    totals: dict[str, int] = {k: sum(row[k] for row in rows) for k in _COUNT_KEYS}
    totals_row = 19
    ws.cell(totals_row, 4,  totals["this_city_male"])
    ws.cell(totals_row, 5,  totals["this_city_female"])
    ws.cell(totals_row, 7,  totals["other_city_male"])
    ws.cell(totals_row, 8,  totals["other_city_female"])
    ws.cell(totals_row, 10, totals["other_province_male"])
    ws.cell(totals_row, 11, totals["other_province_female"])
    ws.cell(totals_row, 13, totals["foreign_male"])
    if ws.max_column >= 14:
        ws.cell(totals_row, 14, totals["foreign_female"])

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"DTA3_{_safe_filename_part(lgu_title)}.xlsx"
    return buf.getvalue(), fname


# ── Keep old names as aliases so nothing else breaks ──────────────────────────
def export_day_tour_workbook(*, lgu_id: int, report_date: date | None = None) -> tuple[bytes, str]:
    return export_combined_workbook(lgu_id=lgu_id, report_date=report_date)


def export_overnight_workbook(*, lgu_id: int, report_date: date | None = None) -> tuple[bytes, str]:
    return export_combined_workbook(lgu_id=lgu_id, report_date=report_date)
